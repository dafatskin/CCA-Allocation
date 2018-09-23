#############
#data module#
#############

import openpyxl

import timeit


def configure_variables(config_file):
    """
    Returns all the variables stored in the config file
    {data type:{data}}
    """

    print("configuring variables...")
    
    dic = {} #will return to config_vars

    try: #test (and open) file
        fp = open(config_file, "r")
    except IOError:
        print(e)
        print("The config file has been moved or renamed.")
        print("Please return it back to this directory or rename it to 'Config file'.")
    data = fp.readlines()
    fp.close()

    section = "None" #different section = different data format

    to_add = {}
    for line in data: #each line
        if line[:2] == "--" and line[8:].strip("\n") != section: #new section?
            section = line[8:].strip("\n")
            to_add = {}
        elif line[:2] == "--" and line[8:].strip("\n") == section: #end of section?
            dic[section[:-2]] = to_add
            section = "None"
            to_add = {}
        else:
            if section == "data formats--": #section specifying data form
                elements = line.strip("\n").split(":")
                elements.append(elements[1].split(","))
                del elements[1]
                areas = []
                
                for i in range(len(elements[1])): #for each container
                    var = elements[1][i].split("-")
                    areas.append({"slice_coords":[var[0], var[1]], "ID_header":var[2]})
                    
                elements.append(areas)
                del elements[1]

                to_add[elements[0]] = elements[1]
            elif section == "file names--": #section specifying file names
                elements = line.strip("\n").split(":")
                to_add[elements[0]] = elements[1]
            elif section == "scoring details--": #section specifying scoring details
                elements = line.strip("\n").split(":")
                elements.append(elements[1].split(","))
                del elements[1]
                details = {}

                for i in range(len(elements[1])): #for each detail
                    if elements[1][i] == "on" or elements[1][i] == "off":
                        details["rankscoring"] = elements[1][i]
                    else:
                        var = elements[1][i].split("-")
                        lst = ["Height", "Weight", "30m", "IPU", "SBJ", "1km"]
                        details[lst[i]] = {"default":var[0], "other":var[1], "criteria":var[2]}

                elements.append(details)
                del elements[1]

                to_add[elements[0]] = elements[1]
            elif section == "reassignment variables--":
                elements = line.strip("\n").split(":")
                to_add[elements[0]] = elements[1]
            elif section == "None":
                pass
            else:
                print("Error occured on line 50: section '{}' not found".format(section))
                
    return dic
            
            
def extract_data(file, config_vars):
    """
    Reads all data from the file specified in config file, and writes the data to a nested dictionary
    {sheet:{ID:{data}}
    Returns this dictionary
    """

    print("extracting data...")
    
    dic = {} #will return to master_list
    
    try: #test (and open) file
        fp = open(file, "r")
    except IOError as e:
        print(e)
        print("File: '{}' not found. Please check the config file.".format(file))
        

    wb = openpyxl.load_workbook(file) #open workbook
    sheet_names = wb.sheetnames #get names of all sheets

    for name in sheet_names: #for every sheet in the workbook
        sheet = wb[name] #define the worksheet
        sheet_list = []

        areas = config_vars["data formats"][name]

        sheet_list = extract_sheet(file, name, areas) #see extract_sheet

        dic[name] = sheet_list #define sheet as a list of data containers

    fp.close()

    return dic


def extract_sheet(file, sheetname, areas):
    """
    Extracts an individual sheet, used in extract_data
    """

    lst = [] #will return to sheet_data

    try: #test (and open) file
        fp = open(file, "r")
    except IOError as e:
        print(e)
        print("File: '{}' not found. Please check the config file.".format(file))

    wb = openpyxl.load_workbook(file) #open workbook

    try: #test (and open) spreadsheet
        ws = wb[sheetname]
    except KeyError as e:
        print(e)
        print("Sheet: '{}' not found. Please check the config file.".format(sheetname))

    for i in areas: #for each area
        area = ws[i["slice_coords"][0]:i["slice_coords"][1]]
        area_dic = {}
        ID_value = ""

        for row in area: #for each row in area
            row_dic = {}

            for cell in row: #for each cell in row
                col_letter = cell.column #this be column of cell
                header = ws[col_letter + i["slice_coords"][0][1:]].value #this be header value of cell
                if header == i["ID_header"]: #if its the ID column
                    ID_value = cell.value #get the ID value
                else:
                    row_dic[header] = cell.value #define column of ID as value

            area_dic[ID_value] = row_dic #define ID of area as column

        lst.append(area_dic) #add to list of areas

    fp.close()    
    return lst
            

def data_to_LOS(dic):
    """
    Returns a list of all students in directory
    [name]
    """

    final_lst = []

    dic_classlist = dic["classlist"][0] #relevant sheet

    for key, value in classlist.items(): #name:data
        final_lst.append(key)

    del final_lst[0]

    return final_lst


def data_to_LOC(dic):
    """
    Returns a dictionary of core cca choices of each student
    {rank of choice:{student:cca}}
    """

    final_dic = {} #will return to list_of_firsts
    
    dic_choices = dic["choices"][0] #the relevant sheet

    pdic = {"LO1":"CORE1", "LO2":"CORE2", "LO3":"CORE3", "LO4":"CORE4", "LO5":"CORE5", "LO6":"CORE6", "LO7":"CORE7", "LO8":"CORE8", "LO9":"CORE9"}
    qdic = {}
    for key, value in pdic.items(): #for each rank:name of rank
        for NRIC, choices in dic_choices.items(): #for each student:choices
            choice = ""
            if choices[value] == "01SCOUT": #these 2 values have changes later on. Standardising
                choice = "O1"
            elif choices[value] == "02SCOUT":
                choice = "O2"
            else:
                choice = choices[value]
            qdic[NRIC] = choice
        final_dic[key] = qdic
        qdic = {}

    return final_dic


def data_to_merit(dic):
    """
    Returns a dictionary of merit cca choices of each student
    {student:merit cca}
    """

    final_dic = {}

    dic_choices = dic["choices"][0] #relevant sheet

    for NRIC, choices in dic_choices.items():
        final_dic[NRIC] = choices["MERIT1"] #just take first choice; no limit for merit CCAs

    del final_dic["NRIC"]

    return final_dic


def data_to_MEP(dic):
    """
    Returns a list of MEP students
    [name]
    """

    final_lst = []

    dic_MEP = dic["MEP"][0] #relevant sheet

    for key, value in dic_MEP.items():
        final_lst.append(key) #just append the name

    del final_lst[0]
  
    return final_lst
    
    
def data_to_DSA(dic):
    """
    Returns a dictionary of DSA students
    {name:CCA}
    """

    final_dic = {} #will return to DSA_students

    dic_DSA = dic["DSA"][0] #the relevant sheet

    for key, value in dic_DSA.items():
        final_dic[key] = value["Sports"]

    del final_dic["Name"]
    
    return final_dic


def data_to_quota(dic):
    """
    Returns a dictionary of quota of each CCA
    {CCA type:{CCA:quota}}
    """

    final_dic = {} #will return to CCA_quota

    dic_quota = dic["ccaquota"] #the relevant sheet

    for dic in dic_quota: #SPORTS, UNIFORMED GROUPS, etc.
        groupname = ""
        groupdic = {}
        for key, value in dic.items(): #SPORTS: {}
            if value[None] == None:
                final_dic[groupname] = groupdic
                groupname = key
                groupdic = {}
            else:
                groupdic[key] = value["QUOTA"]
        final_dic[groupname] = groupdic

    del final_dic[""]
    
    return final_dic


def data_to_psychomotor(dic):
    """
    Returns a dictionary of psychomotor details of each student
    {name:{details}}
    """

    final_dic = {} #will return to psychomotor

    dic_psymo = dic["psychomotor"][0] #the relevant sheet

    for key, value in dic_psymo.items():
        del value["AGE"]
        final_dic[key] = value

    del final_dic["Name"]

    return final_dic


def data_to_CCA(dic, CCA):
    """
    Returns a dictionary of ranking details of each CCA
    {name:{placeholder:rank}
    """

    final_dic = {}

    dic_CCA = dic[CCA][0] #the cca sheet

    for key, value in dic_CCA.items():
        try: #delete all the useless info
            del value["Class"]
        except KeyError:
            del value["CLASS"]
        try:
            del value["Category"]
        except:
            pass
        final_dic[key] = value

    try:
        del final_dic["Name"]
    except KeyError:
        pass

    return final_dic


def data_to_nameCat(LOC, quota, rank, CCA):
    """
    Returns a dictionary of the category of a CCA
    """

    final_dic = {}
    
    dic_quota = quota.dic #dictionary

    cat = ""

    for category, dic_CCAs in dic_quota.items(): #for each category
        for cca, quota in dic_CCAs.items(): #for each cca
            if cca == CCA:
                cat = category #variable = category of cca
            else:
                pass

    CCA_LOC = {} #reverse LOC
    for name, cca in LOC.dic[rank].items():
        try:
            lst = CCA_LOC[cca]
            lst.append(name)
            CCA_LOC[cca] = lst
        except KeyError:
            CCA_LOC[cca] = [name]

    try:
        for name in CCA_LOC[CCA]:
            final_dic[name] = cat #name:category
    except KeyError:
        pass

    try:
        del final_dic["Name"]
    except KeyError:
        pass
    
    return final_dic


def data_to_nameClass(master_list):
    """
    Returns a dictionary of students' classes
    {name:class}
    """

    final_dic = {}

    dic_classlist = master_list["classlist"][0] #relevant sheet

    for name, data in dic_classlist.items(): 
        final_dic[name] = data["CLASS"]

    del final_dic["NAME"]
    return final_dic
    
